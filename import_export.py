# import_export.py
"""
Import / Export helpers for Game Manager GUI
Handles CSV, Excel, JSON, SQLite, TXT files with URL normalization
Automatically fixes IGDB image URLs (\\images.igdb.com → https://images.igdb.com)
"""

from __future__ import annotations
import os
import json
import csv
import sqlite3
import hashlib
import html
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Tuple, List, Dict, Optional, Any
import tempfile
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Optional libraries - import if available
pd = None
load_workbook = None
REPORTLAB_AVAILABLE = False
FPDF_AVAILABLE = False

# Try pandas for Excel import
try:
    import pandas as pd
except ImportError:
    pd = None

# Try openpyxl for Excel import
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# Try reportlab for PDF export
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import mm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Try fpdf for PDF export (fallback)
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Import sanitize helper from your project
try:
    from utils_sanitize import sanitize_original_title, load_repack_list
except ImportError:
    # Fallback if not available
    def sanitize_original_title(s: str) -> dict:
        return {"base_title": s, "version": "", "repack": "", "notes": "", "modes": []}
    
    def load_repack_list():
        return []

# Default cache directory for storing images
DEFAULT_CACHE_BASE = os.path.join(tempfile.gettempdir(), "game_manager_cache")
os.makedirs(DEFAULT_CACHE_BASE, exist_ok=True)



# Add these color constants for consistent theming
COLOR_THEME = {
    "primary": "#2c3e50",
    "secondary": "#3498db", 
    "success": "#27ae60",
    "warning": "#f39c12",
    "danger": "#e74c3c",
    "light": "#ecf0f1",
    "dark": "#34495e",
    "info": "#3498db",
    "background": "#f8f9fa",
    "border": "#dee2e6"
}

def _truncate_to_two_lines(text: str, max_width_pts: float, font_size: float) -> str:
    """
    Truncate text to fit within approximately two lines in PDF.
    
    Args:
        text: Text to truncate
        max_width_pts: Maximum width in points
        font_size: Font size in points
    
    Returns:
        Truncated text
    """
    if not text:
        return ""
    
    # Approximate average character width (Helvetica)
    avg_char_width = font_size * 0.55
    chars_per_line = max(10, int(max_width_pts / avg_char_width))
    max_chars = chars_per_line * 2  # Two lines
    
    if len(text) <= max_chars:
        return text
    
    # Truncate and add ellipsis
    return text[:max_chars - 3] + "..."

def empty_game(title: str = "") -> Dict[str, Any]:
    """Return a blank game dictionary with all required fields."""
    return {
        "title": title,
        "app_id": "",
        "release_date": "",
        "developer": "",
        "publisher": "",
        "genres": "",
        "description": "",
        "cover_url": "",
        "trailer_webm": "",
        "screenshots": [],
        "image_cache_paths": [],
        "microtrailer_cache_path": [],
        "shortcut_links": "",
        "steam_link": "",
        "steamdb_link": "",
        "pcgw_link": "",
        "igdb_link": "",
        "save_location": "",
        "savegame_location": [],
        "game_drive": "",
        "scene_repack": "",
        "game_modes": "",
        "original_title": "",
        "original_title_base": "",
        "original_title_version": "",
        "original_notes": "",
        "patch_version": "",
        "player_perspective": "",
        "themes": "",
        "igdb_id": "",
        "played": False,
        "trailers": [],  # Add this line
    }

# Map various column names to standard field names
_HEADER_MAP = {
    "title": "title", "steam id": "app_id", "appid": "app_id",
    "release": "release_date", "release date": "release_date",
    "developer": "developer", "publisher": "publisher", "genres": "genres",
    "trailer": "trailer_webm", "screenshots": "screenshots",
    "steam": "steam_link", "steamdb": "steamdb_link", 
    "pcgamingwiki": "pcgw_link", "pcgw": "pcgw_link",
    "igdb": "igdb_link", "game drive": "game_drive", "game_drive": "game_drive",
    "scene/repack": "scene_repack", "scene": "scene_repack",
    "game modes": "game_modes", "modes": "game_modes",
    "original title": "original_title", "original": "original_title",
    "patch/version": "patch_version", "patch": "patch_version", "version": "patch_version",
    "played": "played", "save location": "save_location", "savegame": "save_location",
    "player perspective": "player_perspective", "player_perspective": "player_perspective",
    "themes": "themes", "igdb id": "igdb_id", "igdb_id": "igdb_id",
    "shortcut_links": "shortcut_links", "shortcut link": "shortcut_links",
    "image_cache_paths": "image_cache_paths", "image_cache_path": "image_cache_paths",
    "savegame_location": "savegame_location", "savegame_locations": "savegame_location",
    "trailers": "trailers",
    "microtrailers_extra": "microtrailers_extra",
    "microtrailer_cache_path": "microtrailer_cache_path","microtrailer_cache_path": "microtrailer_cache_path",
}

def normalize_headers(headers: List[str]) -> List[str]:
    """Convert column headers to standard field names."""
    return [_HEADER_MAP.get(h.strip().lower(), h.strip().lower()) for h in headers]

def _normalize_url(url: str) -> str:
    """
    Fix common URL issues:
    1. Convert \\images.igdb.com → https://images.igdb.com
    2. Convert //images.igdb.com → https://images.igdb.com  
    3. Ensure all URLs have proper https:// prefix
    4. Remove trailing slashes
    """
    if not url or not isinstance(url, str):
        return ""
    
    url = url.strip()
    
    # FIX: Convert backslashes to forward slashes (\\images → //images)
    url = url.replace("\\", "/")
    
    # Handle protocol-relative URLs (//images.igdb.com)
    if url.startswith("//"):
        url = "https:" + url
    
    # Ensure proper protocol for IGDB URLs
    if "images.igdb.com" in url and not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    
    # Remove trailing slashes
    url = url.rstrip("/")
    
    return url

def _enhance_igdb_images(game: Dict) -> Dict:
    """
    Convert IGDB image URLs to 720p quality and fix URL format.
    Example: //images.igdb.com/.../t_thumb/... → https://images.igdb.com/.../t_720p/...
    """
    # Fix cover URL
    cover_url = game.get("cover_url", "")
    if cover_url:
        # First normalize the URL (fix \\ or // issues)
        cover_url = _normalize_url(cover_url)
        
        # Convert to 720p if it's an IGDB URL
        if "images.igdb.com" in cover_url:
            # List of smaller sizes to upgrade to 720p
            small_sizes = ["t_thumb", "t_cover_small", "t_cover_big", 
                          "t_logo_med", "t_screenshot_med", "t_screenshot_big"]
            for size in small_sizes:
                if f"/{size}/" in cover_url:
                    cover_url = cover_url.replace(f"/{size}/", "/t_720p/")
                    break
        
        game["cover_url"] = cover_url
    
    # Fix screenshot URLs
    screenshots = game.get("screenshots", [])
    if screenshots:
        fixed_screenshots = []
        for screenshot in screenshots:
            if screenshot:
                # Normalize the URL first
                screenshot = _normalize_url(screenshot)
                
                # Convert to 720p if it's an IGDB URL
                if "images.igdb.com" in screenshot:
                    small_sizes = ["t_thumb", "t_cover_small", "t_cover_big", 
                                  "t_logo_med", "t_screenshot_med", "t_screenshot_big"]
                    for size in small_sizes:
                        if f"/{size}/" in screenshot:
                            screenshot = screenshot.replace(f"/{size}/", "/t_720p/")
                            break
                
                fixed_screenshots.append(screenshot)
        
        game["screenshots"] = fixed_screenshots
    
    return game

# -----------------------------------------------------------------
# IMPORT FUNCTIONS
# -----------------------------------------------------------------

def import_excel(path: str) -> Tuple[List[Dict], Optional[str]]:
    """Read games from Excel file (.xlsx or .xls)."""
    try:
        if pd is not None:
            # Use pandas if available
            df = pd.read_excel(path, dtype=str).fillna("")
            raw_headers = list(df.columns)
            headers = normalize_headers(raw_headers)
            new_rows = []
            
            for _, row in df.iterrows():
                game = empty_game()
                for h_raw, h in zip(raw_headers, headers):
                    val = row[h_raw]
                    if pd.isna(val) or val is None or str(val).strip() == "":
                        continue
                    
                    val_str = str(val).strip()
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val_str.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    else:
                        game[h] = val_str
                
                # Set title from original_title if title is empty
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game.get("original_title", ""))
                    game["original_title_base"] = san.get("base_title", "")
                    game["original_title_version"] = san.get("version", "")
                    game["scene_repack"] = game.get("scene_repack") or san.get("repack", "")
                    game["original_notes"] = san.get("notes", "")
                    game["game_modes"] = game.get("game_modes") or ", ".join(san.get("modes", []))
                    game["title"] = san.get("base_title") or game["original_title"]
                
                # Fix IGDB image URLs
                game = _enhance_igdb_images(game)
                new_rows.append(game)
            
            return new_rows, None
            
        elif load_workbook is not None:
            # Use openpyxl if available
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.rows)
            
            if not rows:
                return [], "Excel file empty"
            
            raw_headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
            headers = normalize_headers(raw_headers)
            new_rows = []
            
            for r in rows[1:]:
                game = empty_game()
                for i, cell in enumerate(r):
                    if i >= len(headers):
                        break
                    h = headers[i]
                    val = "" if cell.value is None else str(cell.value).strip()
                    if not val:
                        continue
                    
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val.split("|") if s.strip()]
                    else:
                        game[h] = val
                
                # Set title from original_title if title is empty
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game.get("original_title", ""))
                    game["original_title_base"] = san.get("base_title", "")
                    game["original_title_version"] = san.get("version", "")
                    game["scene_repack"] = game.get("scene_repack") or san.get("repack", "")
                    game["original_notes"] = san.get("notes", "")
                    game["game_modes"] = game.get("game_modes") or ", ".join(san.get("modes", []))
                    game["title"] = san.get("base_title") or game["original_title"]
                
                # Fix IGDB image URLs
                game = _enhance_igdb_images(game)
                new_rows.append(game)
            
            return new_rows, None
            
        else:
            return [], "No Excel reader available (install pandas or openpyxl)"
            
    except Exception as e:
        return [], str(e)

def import_csv(path: str) -> Tuple[List[Dict], Optional[str]]:
    """Read games from CSV file."""
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            raw_headers = next(reader, [])
            headers = normalize_headers(raw_headers)
            new_rows = []
            
            for row in reader:
                if not row:
                    continue
                
                game = empty_game()
                for h, val in zip(headers, row):
                    if not val:
                        continue
                    
                    val_str = str(val).strip()
                    if h == "screenshots":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "savegame_location":
                        game["savegame_location"] = [s.strip() for s in val_str.split("|") if s.strip()]
                    elif h == "played":
                        game[h] = val_str.lower() in ("yes", "y", "true", "1", "checked")
                    elif h == "image_cache_paths":
                        game[h] = [s.strip() for s in val_str.split("|") if s.strip()]
                    else:
                        game[h] = val_str
                
                # Set title from original_title if title is empty
                if not game.get("title") and game.get("original_title"):
                    san = sanitize_original_title(game.get("original_title", ""))
                    game["original_title_base"] = san.get("base_title", "")
                    game["original_title_version"] = san.get("version", "")
                    game["scene_repack"] = game.get("scene_repack") or san.get("repack", "")
                    game["original_notes"] = san.get("notes", "")
                    game["game_modes"] = game.get("game_modes") or ", ".join(san.get("modes", []))
                    game["title"] = san.get("base_title") or game["original_title"]
                
                # Fix IGDB image URLs
                game = _enhance_igdb_images(game)
                new_rows.append(game)
        
        return new_rows, None
        
    except Exception as e:
        return [], str(e)

def import_txt(path: str) -> Tuple[List[Dict], Optional[str]]:
    """Read game titles from text file (one per line)."""
    try:
        new_rows = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                original = line.strip()
                if not original:
                    continue
                
                game = empty_game()
                game["original_title"] = original
                san = sanitize_original_title(original)
                game["original_title_base"] = san.get("base_title", "")
                game["original_title_version"] = san.get("version", "")
                game["scene_repack"] = san.get("repack", "")
                game["original_notes"] = san.get("notes", "")
                game["game_modes"] = ", ".join(san.get("modes", []))
                game["title"] = san.get("base_title") or original
                
                # Fix IGDB image URLs (though TXT files usually don't have them)
                game = _enhance_igdb_images(game)
                new_rows.append(game)
        
        return new_rows, None
        
    except Exception as e:
        return [], str(e)

# -----------------------------------------------------------------
# JSON FUNCTIONS
# -----------------------------------------------------------------

def save_to_json(path: str, games: List[Dict]) -> Optional[str]:
    """Save games to JSON file."""
    try:
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        
        # Prepare data for JSON (ensure lists are proper Python lists)
        out = []
        for g in games:
            gg = dict(g)
            # Fix IGDB image URLs before saving
            gg = _enhance_igdb_images(gg)
            
            gg["screenshots"] = list(gg.get("screenshots") or [])
            gg["image_cache_paths"] = list(gg.get("image_cache_paths") or [])
            gg["savegame_location"] = list(gg.get("savegame_location") or [])
            out.append(gg)
        
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        
        return None
        
    except Exception as e:
        return str(e)

def load_from_json(path: str) -> Tuple[List[Dict], Optional[str]]:
    """Load games from JSON file."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if not isinstance(data, list):
            return [], "Invalid JSON DB format (expected list)"
        
        out = []
        for r in data:
            if not isinstance(r, dict):
                continue
            
            g = empty_game()
            g.update(r)
            
            # Ensure fields are proper lists
            g["screenshots"] = list(g.get("screenshots") or [])
            g["image_cache_paths"] = list(g.get("image_cache_paths") or [])
            
            if isinstance(g.get("savegame_location"), str):
                g["savegame_location"] = [s.strip() for s in g.get("savegame_location").split("|") if s.strip()]
            else:
                g["savegame_location"] = list(g.get("savegame_location") or [])
            
            # Fix IGDB image URLs
            g = _enhance_igdb_images(g)
            out.append(g)
        
        return out, None
        
    except Exception as e:
        return [], str(e)

# -----------------------------------------------------------------
# SQLITE FUNCTIONS
# -----------------------------------------------------------------

def save_to_sqlite(db_path: str, games: List[Dict]) -> Optional[str]:
    """Save games to SQLite database."""
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        
        # Create table if it doesn't exist
        c.execute("""
            CREATE TABLE IF NOT EXISTS games (
                id TEXT PRIMARY KEY,
                title TEXT,
                app_id TEXT,
                release_date TEXT,
                developer TEXT,
                publisher TEXT,
                genres TEXT,
                description TEXT,
                cover_url TEXT,
                trailer_webm TEXT,
                screenshots TEXT,
                image_cache_paths TEXT,
                shortcut_links TEXT,
                steam_link TEXT,
                steamdb_link TEXT,
                pcgw_link TEXT,
                igdb_link TEXT,
                save_location TEXT,
                savegame_location TEXT,
                game_drive TEXT,
                scene_repack TEXT,
                game_modes TEXT,
                original_title TEXT,
                original_title_base TEXT,
                original_title_version TEXT,
                original_notes TEXT,
                patch_version TEXT,
                player_perspective TEXT,
                themes TEXT,
                igdb_id TEXT,
                played INTEGER
            )
        """)
        
        for g in games:
            # Fix IGDB image URLs before saving
            g = _enhance_igdb_images(g)
            
            app_id = str(g.get("app_id", "")).strip()
            # Create unique ID from app_id or title
            gid = app_id if app_id and app_id != "Not Found" else (g.get("title", "").strip() or None)
            
            if not gid:
                gid = hashlib.sha256(json.dumps(g, sort_keys=True).encode("utf-8")).hexdigest()
            
            # Convert lists to JSON strings for SQLite
            shots_json = json.dumps(g.get("screenshots", []), ensure_ascii=False)
            cache_json = json.dumps(g.get("image_cache_paths", []), ensure_ascii=False)
            save_json = json.dumps(g.get("savegame_location", []), ensure_ascii=False)
            played_int = 1 if g.get("played", False) else 0
            shortcut = g.get("shortcut_links") or ""
            
            # Insert or replace game
            c.execute("""
                INSERT OR REPLACE INTO games VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """, (
                gid, g.get("title", ""), app_id, g.get("release_date", ""), 
                g.get("developer", ""), g.get("publisher", ""), g.get("genres", ""),
                g.get("description", ""), g.get("cover_url", ""), g.get("trailer_webm", ""),
                shots_json, cache_json, shortcut, g.get("steam_link", ""),
                g.get("steamdb_link", ""), g.get("pcgw_link", ""), g.get("igdb_link", ""),
                g.get("save_location", ""), save_json, g.get("game_drive", ""),
                g.get("scene_repack", ""), g.get("game_modes", ""), g.get("original_title", ""),
                g.get("original_title_base", ""), g.get("original_title_version", ""),
                g.get("original_notes", ""), g.get("patch_version", ""),
                g.get("player_perspective", ""), g.get("themes", ""), g.get("igdb_id", ""),
                played_int
            ))
        
        conn.commit()
        conn.close()
        return None
        
    except Exception as e:
        return str(e)

def load_from_sqlite(db_path: str) -> Tuple[List[Dict], Optional[str]]:
    """Load games from SQLite database."""
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        
        c.execute("""
            SELECT title, app_id, release_date, developer, publisher, genres, description,
                   cover_url, trailer_webm, screenshots, image_cache_paths, shortcut_links,
                   steam_link, steamdb_link, pcgw_link, igdb_link, save_location, savegame_location,
                   game_drive, scene_repack, game_modes, original_title, original_title_base,
                   original_title_version, original_notes, patch_version, player_perspective, themes,
                   igdb_id, played
            FROM games
        """)
        
        rows = c.fetchall()
        conn.close()
        
        games = []
        for r in rows:
            # Parse JSON strings back to lists
            try:
                shots = json.loads(r[9] or "[]")
            except Exception:
                shots = []
            
            try:
                cache_paths = json.loads(r[10] or "[]")
            except Exception:
                cache_paths = []
            
            try:
                save_list = json.loads(r[17] or "[]")
            except Exception:
                save_list = []
            
            played_bool = bool(r[29]) if len(r) > 29 else False
            
            game = {
                "title": r[0] or "", "app_id": r[1] or "", "release_date": r[2] or "",
                "developer": r[3] or "", "publisher": r[4] or "", "genres": r[5] or "",
                "description": r[6] or "", "cover_url": r[7] or "", "trailer_webm": r[8] or "",
                "screenshots": shots, "image_cache_paths": cache_paths, "shortcut_links": r[11] or "",
                "steam_link": r[12] or "", "steamdb_link": r[13] or "", "pcgw_link": r[14] or "",
                "igdb_link": r[15] or "", "save_location": r[16] or "", "savegame_location": save_list,
                "game_drive": r[18] or "", "scene_repack": r[19] or "", "game_modes": r[20] or "",
                "original_title": r[21] or "", "original_title_base": r[22] or "",
                "original_title_version": r[23] or "", "original_notes": r[24] or "",
                "patch_version": r[25] or "", "player_perspective": r[26] or "",
                "themes": r[27] or "", "igdb_id": r[28] or "", "played": played_bool
            }
            
            # Fix IGDB image URLs
            game = _enhance_igdb_images(game)
            games.append(game)
        
        return games, None
        
    except Exception as e:
        return [], str(e)

# -----------------------------------------------------------------
# CACHE FUNCTIONS
# -----------------------------------------------------------------

def game_cache_dir(game: Dict, cache_base: Optional[str] = None) -> str:
    """Get cache directory for a specific game."""
    base = cache_base or DEFAULT_CACHE_BASE
    basep = Path(base)
    appid = str(game.get("app_id") or "").strip()
    
    # Use app_id if available, otherwise hash of title
    if appid:
        sub = f"game_{appid}"
    else:
        key = (game.get("title") or "") + "|" + (game.get("original_title") or "")
        h = hashlib.sha256(key.encode("utf-8")).hexdigest()[:12]
        sub = f"game_{h}"
    
    d = basep / sub
    d.mkdir(parents=True, exist_ok=True)
    return str(d)

def prune_game_cache_dir(game: Dict, keep: int = 8, cache_base: Optional[str] = None) -> None:
    """Keep only the newest files in cache directory."""
    d = game_cache_dir(game, cache_base=cache_base)
    p = Path(d)
    
    if not p.exists():
        return
    
    # Sort files by modification time (newest first)
    files = sorted([f for f in p.iterdir() if f.is_file()], 
                   key=lambda f: f.stat().st_mtime, reverse=True)
    
    # Delete old files beyond 'keep' count
    for old in files[keep:]:
        try:
            old.unlink()
        except Exception:
            pass

def save_image_bytes(game: Dict, url: str, data: bytes, cache_base: Optional[str] = None) -> Optional[str]:
    """Save image to cache and return file path."""
    try:
        d = game_cache_dir(game, cache_base=cache_base)
        # Create hash from URL for filename
        h = hashlib.sha256(url.encode("utf-8")).hexdigest()
        fname = f"{h}.bin"
        path = Path(d) / fname
        
        with open(path, "wb") as fh:
            fh.write(data)
        
        # Clean up old files
        prune_game_cache_dir(game, keep=8, cache_base=cache_base)
        return str(path)
        
    except Exception:
        return None

# -----------------------------------------------------------------
# MERGE FUNCTION
# -----------------------------------------------------------------

def merge_imported_rows(existing_games: List[Dict], imported_rows: List[Dict], 
                        prefer_imported: bool = True) -> List[Dict]:
    """
    Merge imported games into existing list.
    Matches games by app_id first, then by title.
    If prefer_imported=True, imported data overwrites existing.
    """
    # Create lookup dictionaries
    by_app = {}
    by_title = {}
    
    for g in existing_games:
        aid = str(g.get("app_id") or "").strip()
        if aid:
            by_app[aid] = g
        t = g.get("title")
        if t:
            by_title[t] = g
    
    for imp in imported_rows:
        # Fix IGDB image URLs in imported data
        imp = _enhance_igdb_images(imp)
        
        aid = str(imp.get("app_id") or "").strip()
        matched = None
        
        # Try to match by app_id first
        if aid and aid in by_app:
            matched = by_app[aid]
        else:
            # Try to match by title
            t = imp.get("title")
            if t and t in by_title:
                matched = by_title[t]
        
        if matched:
            # Merge data
            for k, v in imp.items():
                if v is None or v == "":
                    continue
                if prefer_imported:
                    matched[k] = v
                else:
                    if not matched.get(k):
                        matched[k] = v
        else:
            # Add as new game
            existing_games.append(imp)
    
    return existing_games

# -----------------------------------------------------------------
# EXPORT FUNCTIONS
# -----------------------------------------------------------------

# Add to imports section at the top
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
import html
from pathlib import Path
import webbrowser

# Add these color constants for consistent theming
COLOR_THEME = {
    "primary": "#2c3e50",
    "secondary": "#3498db",
    "success": "#27ae60",
    "warning": "#f39c12",
    "danger": "#e74c3c",
    "light": "#ecf0f1",
    "dark": "#34495e",
    "info": "#3498db"
}

def export_games_to_pdf(path: str, games: List[Dict], title: Optional[str] = None) -> Optional[str]:
    """
    Export games to a beautifully formatted PDF file with enhanced styling.
    
    Args:
        path: Path to save the PDF file
        games: List of game dictionaries to export
        title: Title for the PDF document
    
    Returns:
        Error message or None if successful
    """
    try:
        if REPORTLAB_AVAILABLE:
            return _export_games_to_pdf_reportlab(path, games, title)
        elif FPDF_AVAILABLE:
            return _export_games_to_pdf_fpdf(path, games, title)
        else:
            return "No PDF library available. Install reportlab or fpdf."
    except Exception as e:
        return f"Error exporting to PDF: {str(e)}"

def _export_games_to_pdf_reportlab(path: str, games: List[Dict], title: Optional[str] = None) -> Optional[str]:
    """PDF export using ReportLab matching download report style."""
    try:
        print(f"Starting ReportLab PDF export to: {path}")
        print(f"Number of games to process: {len(games)}")
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import mm, inch
        import time
        
        # Document setup with VERY NARROW margins
        page_size = landscape(A4)
        print(f"Page size: {page_size}, Landscape A4")
        
        doc = SimpleDocTemplate(
            path,
            pagesize=page_size,
            leftMargin=5*mm,      # Reduced from 8mm
            rightMargin=5*mm,     # Reduced from 8mm
            topMargin=8*mm,       # Reduced from 10mm
            bottomMargin=8*mm,    # Reduced from 10mm
            title=title or "Game Collection Report"
        )
        
        # Calculate usable width in points
        page_width_pts, page_height_pts = page_size
        usable_width = page_width_pts - (5*mm * 2)  # 5mm left + 5mm right
        
        print(f"Usable width: {usable_width} pts")
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Title style (matching download report)
        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Title"],
            fontSize=12,
            textColor=colors.HexColor(COLOR_THEME["primary"]),
            alignment=1,
            spaceAfter=3,
            fontName="Helvetica-Bold"
        )
        
        # Subtitle style
        subtitle_style = ParagraphStyle(
            "SubtitleStyle",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor("#666666"),
            alignment=1,
            spaceAfter=8
        )
        
        # Statistics header style
        stats_header_style = ParagraphStyle(
            "StatsHeaderStyle",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.white,
            fontName="Helvetica-Bold",
            alignment=1,
            leading=9
        )
        
        # Statistics value style
        stats_value_style = ParagraphStyle(
            "StatsValueStyle",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            alignment=1,
            leading=12
        )
        
        # Statistics label style
        stats_label_style = ParagraphStyle(
            "StatsLabelStyle",
            parent=styles["Normal"],
            fontSize=6,
            textColor=colors.HexColor("#666666"),
            alignment=1,
            leading=7
        )
        
        # Header style for tables
        header_style = ParagraphStyle(
            "HeaderStyle",
            parent=styles["Normal"],
            fontSize=7,
            textColor=colors.black,
            fontName="Helvetica-Bold",
            alignment=1,
            leading=8
        )
        
        # Cell style
        cell_style = ParagraphStyle(
            "CellStyle",
            parent=styles["Normal"],
            fontSize=6,
            leading=7,
            wordWrap='CJK'
        )
        
        # Small cell style for resources
        small_cell_style = ParagraphStyle(
            "SmallCellStyle",
            parent=styles["Normal"],
            fontSize=5,
            leading=6,
            wordWrap='CJK'
        )
        
        # No data style
        no_data_style = ParagraphStyle(
            "NoDataStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor(COLOR_THEME["warning"]),
            alignment=1,
            leading=12
        )
        
        # Story content
        story = []
        
        # Title section (matching download report header)
        timestamp = time.strftime("%Y-%m-%d at %H:%M:%S")
        story.append(Paragraph("Game Collection Report", title_style))
        story.append(Paragraph(f"Generated on {timestamp}", subtitle_style))
        
        # Calculate statistics
        total = len(games)
        played = sum(1 for g in games if g.get("played"))
        remaining = total - played
        success_rate = (played / total * 100) if total > 0 else 0
        
        print(f"Statistics - Total: {total}, Played: {played}, Remaining: {remaining}")
        
        # Statistics table (matching download report stats-grid)
        stats_data = [
            [
                Paragraph("Total Games", stats_header_style),
                Paragraph("Played Games", stats_header_style),
                Paragraph("Remaining Games", stats_header_style)
            ],
            [
                Paragraph(str(total), ParagraphStyle(
                    "TotalStatsStyle",
                    parent=stats_value_style,
                    textColor=colors.HexColor(COLOR_THEME["primary"])
                )),
                Paragraph(str(played), ParagraphStyle(
                    "PlayedStatsStyle",
                    parent=stats_value_style,
                    textColor=colors.HexColor(COLOR_THEME["success"])
                )),
                Paragraph(str(remaining), ParagraphStyle(
                    "RemainingStatsStyle",
                    parent=stats_value_style,
                    textColor=colors.HexColor(COLOR_THEME["warning"])
                ))
            ],
            [
                Paragraph("In collection", stats_label_style),
                Paragraph(f"{success_rate:.1f}% completion", stats_label_style),
                Paragraph("Games to play", stats_label_style)
            ]
        ]
        
        stats_table = Table(stats_data, colWidths=[usable_width/3]*3)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_THEME["primary"])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, 1), 10),
            ('FONTSIZE', (0, 2), (-1, 2), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
            ('TOPPADDING', (0, 1), (-1, 1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 4),
            ('TOPPADDING', (0, 2), (-1, 2), 2),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 8))
        
        # Game list table - ADDED "Theme" column after "Genre"
        story.append(Paragraph("<b>Game List</b>", ParagraphStyle(
            "Heading2Style",
            parent=styles["Heading2"],
            fontSize=9,
            spaceAfter=6
        )))
        
        # Check if there are games to display
        if not games:
            print("No games to display, adding message...")
            # Add a message when there are no games
            story.append(Paragraph("No games in the database. Please import some games first.", no_data_style))
            story.append(Spacer(1, 12))
        else:
            # Prepare game table data - Added "Theme" column
            headers = ["SN", "Played", "Title", "Steam ID", "Genre", "Theme", "Description", "Mode", "Drive", "Original", "Resources"]
            
            # Calculate column widths as percentages of usable_width (matching stats table width)
            # These percentages should sum to 1.0 (100% of usable_width)
            col_percentages = [
                0.015,  # SN: 1.5%
                0.025,  # Played: 2.5%
                0.135,  # Title: 13.5%
                0.050,  # Steam ID: 5.0%
                0.080,  # Genre: 8.0%
                0.065,  # Theme: 6.5%
                0.250,  # Description: 25.0%
                0.055,  # Mode: 5.5%
                0.050,  # Drive: 5.0%
                0.190,  # Original: 19.0%
                0.120,  # Resources: 12.0%
            ]
            
            # Convert percentages to actual widths
            col_widths = [usable_width * p for p in col_percentages]
            
            # Verify the sum is close to usable_width
            total_width = sum(col_widths)
            print(f"Table width: {total_width:.1f} pts, Usable width: {usable_width:.1f} pts, Match: {abs(total_width - usable_width) < 0.1}")
            
            table_data = []
            
            # Create header row with proper Paragraph objects
            header_cells = [Paragraph(h, header_style) for h in headers]
            table_data.append(header_cells)
            
            print(f"Processing {len(games)} games for table...")
            
            # Add game rows
            for idx, game in enumerate(games, 1):
                # Get data with truncation - adjust based on column widths
                title_text = game.get("title", "") or "Untitled"
                steam_id = str(game.get("app_id", "") or "")
                genre = game.get("genres", "") or ""
                theme = game.get("themes", "") or ""  # NEW: Get theme
                desc = (game.get("description", "") or "")[:150] + ("..." if len(game.get("description", "")) > 150 else "")
                mode = game.get("game_modes", "") or ""
                drive = game.get("game_drive", "") or ""
                original = game.get("original_title", "") or ""
                
                # Played status with color
                played_status = "Yes" if game.get("played") else "No"
                played_color = COLOR_THEME["success"] if game.get("played") else COLOR_THEME["warning"]
                
                # RESOURCES: Screenshots and Trailers
                # Get screenshots
                screenshots = list(game.get("screenshots", []) or [])
                cache_paths = list(game.get("image_cache_paths", []) or [])
                all_screenshots = screenshots + cache_paths
                
                # Get trailers
                trailers = []
                trailer_webm = game.get("trailer_webm", "")
                if trailer_webm:
                    trailers.append(trailer_webm)
                
                # Build screenshot links
                ss_links = []
                for i, url in enumerate(all_screenshots[:4], 1):
                    if url:  # Check if url is not empty
                        safe_url = html.escape(url)
                        ss_links.append(f'<a href="{safe_url}">[{i}]</a>')
                
                # Build trailer links  
                tt_links = []
                for i, url in enumerate(trailers[:2], 1):
                    if url:  # Check if url is not empty
                        safe_url = html.escape(url)
                        tt_links.append(f'<a href="{safe_url}">[{i}]</a>')
                
                # Create resources text
                resources_text = ""
                if ss_links:
                    resources_text += f'<b>SS:</b> {" ".join(ss_links)}<br/>'
                if tt_links:
                    resources_text += f'<b>TT:</b> {" ".join(tt_links)}'
                
                if not resources_text:
                    resources_text = "None"
                
                # Calculate approximate characters per column based on width
                # Create row with NEW Theme column
                row = [
                    Paragraph(str(idx), cell_style),
                    Paragraph(f'<font color="{played_color}">{played_status}</font>', cell_style),
                    Paragraph(title_text[:35] + ("..." if len(title_text) > 35 else ""), cell_style),
                    Paragraph(steam_id[:10] + ("..." if len(steam_id) > 10 else ""), cell_style),
                    Paragraph(genre[:25] + ("..." if len(genre) > 25 else ""), cell_style),
                    Paragraph(theme[:20] + ("..." if len(theme) > 20 else ""), cell_style),
                    Paragraph(desc, cell_style),
                    Paragraph(mode[:15] + ("..." if len(mode) > 15 else ""), cell_style),
                    Paragraph(drive[:15] + ("..." if len(drive) > 15 else ""), cell_style),
                    Paragraph(original[:45] + ("..." if len(original) > 45 else ""), cell_style),
                    Paragraph(resources_text, small_cell_style)
                ]
                
                table_data.append(row)
            
            print(f"Created table with {len(table_data)} rows (1 header + {len(games)} data)")
            
            # Create main table - use the same usable_width as stats table
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f2f2f2")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
                
                # Alignment
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'CENTER'),
                ('ALIGN', (7, 0), (7, -1), 'CENTER'),  # Mode column center aligned
                ('ALIGN', (10, 0), (10, -1), 'LEFT'),
                
                # Padding - use minimal padding
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                
                # Vertical alignment
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                
                # Text color for all cells
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                
                # Make sure table uses full width
                ('WIDTH', (0, 0), (-1, -1), usable_width),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 6))
        
        # Footer
        footer_text = f"""
        <para alignment="center">
        <font size="7" color="#666666">
        Report generated by Game Manager • {timestamp} • 
        Total: {total} games • Played: {played} • Remaining: {remaining}
        </font>
        </para>
        """
        story.append(Paragraph(footer_text, subtitle_style))
        
        print("Building PDF document...")
        
        # Build document
        doc.build(story)
        print(f"PDF successfully created at: {path}")
        return None
        
    except Exception as e:
        error_msg = f"ReportLab PDF export error: {str(e)}"
        print(f"ERROR in _export_games_to_pdf_reportlab: {error_msg}")
        import traceback
        traceback.print_exc()
        return error_msg
        
def _export_games_to_pdf_fpdf(path: str, games: List[Dict], title: Optional[str] = None) -> Optional[str]:
    """Fallback PDF export using FPDF."""
    try:
        from fpdf import FPDF
        import time
        
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        # Title
        pdf.set_font("Arial", 'B', 16)
        pdf.set_text_color(44, 62, 80)  # primary color
        pdf.cell(0, 10, title or "Game Manager Export", ln=True, align='C')
        
        # Date
        pdf.set_font("Arial", 'I', 10)
        pdf.set_text_color(52, 73, 94)  # dark color
        pdf.cell(0, 8, f"Generated on {time.strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        
        pdf.ln(5)
        
        # Table headers - Changed "Screenshots" to "Resources"
        headers = ["#", "✓", "Title", "Steam ID", "Genre", "Description", "Mode", "Drive", "Original", "Resources"]
        col_widths = [8, 8, 40, 20, 25, 70, 20, 20, 40, 25]
        
        # Header row
        pdf.set_fill_color(44, 62, 80)  # primary
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 9)
        
        for header, width in zip(headers, col_widths):
            pdf.cell(width, 8, header, border=1, ln=0, align='C', fill=True)
        pdf.ln()
        
        # Game rows
        pdf.set_font("Arial", '', 8)
        
        for idx, game in enumerate(games, 1):
            # Alternate row colors
            fill_color = (255, 255, 255) if idx % 2 == 0 else (248, 249, 250)
            pdf.set_fill_color(*fill_color)
            
            # Get data
            title_text = (game.get("title", "") or "")[:35]
            steam_id = str(game.get("app_id", "") or "")[:10]
            genre = (game.get("genres", "") or "")[:20]
            desc = (game.get("description", "") or "").replace("\n", " ")[:60]
            mode = (game.get("game_modes", "") or "")[:15]
            drive = (game.get("game_drive", "") or "")[:15]
            original = (game.get("original_title", "") or "")[:35]
            
            # Played status
            played = "✓" if game.get("played") else ""
            if played:
                pdf.set_text_color(39, 174, 96)  # success color
            else:
                pdf.set_text_color(100, 100, 100)
            
            pdf.cell(col_widths[0], 6, str(idx), border='LR', ln=0, align='C', fill=True)
            pdf.cell(col_widths[1], 6, played, border='LR', ln=0, align='C', fill=True)
            
            # Reset text color
            pdf.set_text_color(0, 0, 0)
            
            # RESOURCES: Screenshots and Trailers
            # Get screenshots
            screenshots = list(game.get("screenshots", []) or [])
            cache_paths = list(game.get("image_cache_paths", []) or [])
            all_screenshots = screenshots + cache_paths
            
            # Get trailers
            trailers = []
            trailer_webm = game.get("trailer_webm", "")
            if trailer_webm:
                trailers.append(trailer_webm)
            
            # Build resources text
            resources_lines = []
            if all_screenshots:
                ss_text = f"SS: {' '.join(f'[{i+1}]' for i in range(min(3, len(all_screenshots))))}"
                resources_lines.append(ss_text)
            
            if trailers:
                tt_text = f"TT: {' '.join(f'[{i+1}]' for i in range(min(2, len(trailers))))}"
                resources_lines.append(tt_text)
            
            resources_text = "\n".join(resources_lines) if resources_lines else "None"
            
            # Get remaining cells
            cells = [title_text, steam_id, genre, desc, mode, drive, original]
            for i, data in enumerate(cells):
                pdf.cell(col_widths[i+2], 6, data, border='LR', ln=0, align='L', fill=True)
            
            # Resources cell (multi-line)
            pdf.multi_cell(col_widths[9], 3, resources_text, border='LR', ln=0, align='L', fill=True)
            pdf.ln()
            
            # Bottom border
            pdf.set_draw_color(222, 226, 230)  # border color
            pdf.line(10, pdf.get_y(), 280, pdf.get_y())
        
        # Footer
        total = len(games)
        played = sum(1 for g in games if g.get("played"))
        remaining = total - played
        
        pdf.ln(10)
        pdf.set_font("Arial", 'I', 9)
        pdf.set_text_color(52, 73, 94)
        pdf.cell(0, 6, f"Total: {total} | Played: {played} | Remaining: {remaining}", ln=True, align='C')
        
        pdf.output(path)
        return None
        
    except Exception as e:
        return f"FPDF export error: {str(e)}"


def export_games_to_html(path: str, games: List[Dict], title: Optional[str] = None, 
                        open_after: bool = False) -> Optional[str]:
    """
    Export games to an HTML file with download report analytics style.
    
    Args:
        path: Path to save the HTML file
        games: List of game dictionaries to export
        title: Title for the HTML document
        open_after: Open the generated HTML in the default browser
    
    Returns:
        Error message or None if successful
    """
    try:
        from datetime import datetime
        
        # Calculate statistics
        total_games = len(games)
        played_games = sum(1 for g in games if g.get("played"))
        games_with_screenshots = sum(1 for g in games if g.get("screenshots") or g.get("image_cache_paths"))
        games_with_save_info = sum(1 for g in games if g.get("savegame_location"))
        remaining_games = total_games - played_games
        success_rate = (played_games / total_games * 100) if total_games > 0 else 0
        
        timestamp = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
        
        # Generate game rows HTML (compact format, all columns in one row)
        rows_html = []
        for idx, game in enumerate(games, start=1):
            # Get game data
            title_text = html.escape(game.get("title", "") or "Untitled")
            steam_id = html.escape(str(game.get("app_id", "") or ""))
            steam_link = game.get("steam_link") or ""
            genre = html.escape(game.get("genres", "") or "")
            theme = html.escape(game.get("themes", "") or "")  # NEW: Get theme
            description = html.escape(game.get("description", "") or "").replace("\n", " ")
            game_mode = html.escape(game.get("game_modes", "") or "")
            drive = html.escape(game.get("game_drive", "") or "")
            original = html.escape(game.get("original_title", "") or "")
            
            # Played status
            played = "Yes" if game.get("played") else "No"
            played_color = COLOR_THEME["success"] if game.get("played") else COLOR_THEME["warning"]
            
            # Steam link
            if steam_link and steam_id:
                steam_cell = f'<a href="{html.escape(steam_link)}" target="_blank">{steam_id or "Steam"}</a>'
            else:
                steam_cell = steam_id or "N/A"
            
            # RESOURCES: Screenshots and Trailers
            # Get screenshots
            screenshots = list(game.get("screenshots", []) or [])
            cache_paths = list(game.get("image_cache_paths", []) or [])
            all_screenshots = screenshots + cache_paths
            
            # Get trailers (microtrailer first, then trailer_webm)
            trailers = []
            trailer_webm = game.get("trailer_webm", "")
            if trailer_webm:
                trailers.append(trailer_webm)
            
            # Build screenshot links
            ss_links = []
            for i, url in enumerate(all_screenshots[:5], 1):  # Limit to 5
                safe_url = html.escape(url)
                ss_links.append(f'<a href="{safe_url}" target="_blank">[{i}]</a>')
            
            # Build trailer links  
            tt_links = []
            for i, url in enumerate(trailers[:3], 1):  # Limit to 3
                safe_url = html.escape(url)
                tt_links.append(f'<a href="{safe_url}" target="_blank">[{i}]</a>')
            
            # Create resources cell with two lines
            resources_cell = ""
            if ss_links:
                resources_cell += f'<div><strong>SS:</strong> {" ".join(ss_links)}</div>'
            if tt_links:
                resources_cell += f'<div><strong>TT:</strong> {" ".join(tt_links)}</div>'
            
            if not resources_cell:
                resources_cell = f'<span style="color:{COLOR_THEME["warning"]};font-style:italic;">None</span>'
            
            # Create compact row (all columns in one row) - ADDED Theme column
            row_html = f"""
            <tr>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{idx}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6; color: {played_color}; font-weight: bold;">
                    {played}
                </td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;"><strong>{title_text}</strong></td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{steam_cell}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{genre}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{theme}</td> <!-- NEW: Theme column -->
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{description[:80]}{"..." if len(description) > 80 else ""}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6; white-space: nowrap;">{game_mode}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{drive}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6;">{original}</td>
                <td style="padding: 6px; border-bottom: 1px solid #dee2e6; line-height: 1.2;">{resources_cell}</td>
            </tr>
            """
            rows_html.append(row_html)
        
        # Create HTML document matching download report style with NARROWER margins
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>{html.escape(title or "Game Manager Export")}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.4; }} /* Reduced margin from 40px */
        .header {{ background: linear-gradient(to right, {COLOR_THEME['primary']}, #4a6491); color: white; padding: 15px; border-radius: 5px; }} /* Reduced padding */
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 15px 0; }} /* Reduced gap and minmax */
        .stat-card {{ background: #f8f9fa; padding: 12px; border-radius: 5px; border-left: 4px solid {COLOR_THEME['info']}; }} /* Reduced padding */
        .stat-card.success {{ border-left-color: {COLOR_THEME['success']}; }}
        .stat-card.warning {{ border-left-color: {COLOR_THEME['warning']}; }}
        .stat-card.danger {{ border-left-color: {COLOR_THEME['danger']}; }}
        .stat-value {{ font-size: 20px; font-weight: bold; margin: 3px 0; }} /* Reduced font size */
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 12px; }} /* Reduced font size */
        th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }} /* Reduced padding */
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        .failure-section {{ background: #f8d7da; padding: 12px; border-radius: 5px; margin: 15px 0; }}
        .success-section {{ background: #d4edda; padding: 12px; border-radius: 5px; margin: 15px 0; }}
        a {{ color: {COLOR_THEME['info']}; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .footer {{ margin-top: 20px; padding: 12px; background-color: #f8f9fa; border-radius: 5px; text-align: center; font-size: 12px; }} /* Reduced padding and font size */
        .resources-cell div {{ margin-bottom: 3px; }}
        .resources-cell strong {{ color: {COLOR_THEME['primary']}; }}
        .mode-column {{ white-space: nowrap; min-width: 80px; }} /* Added for Mode column */
    </style>
</head>
<body>
    <div class="header">
        <h1 style="margin: 0; font-size: 20px;">Game Collection Report</h1> <!-- Reduced font size -->
        <p style="margin: 5px 0;">Generated on {timestamp}</p>
        <p style="margin: 5px 0;">Total Games: {total_games}</p>
    </div>
    
    <h2 style="font-size: 16px;">Summary</h2> <!-- Reduced font size -->
    <div class="stats-grid">
        <div class="stat-card success">
            <h3 style="margin: 0; font-size: 14px;">Total Games</h3> <!-- Reduced font size -->
            <div class="stat-value">{total_games}</div>
            <p style="margin: 0; font-size: 12px;">In collection</p> <!-- Reduced font size -->
        </div>
        <div class="stat-card">
            <h3 style="margin: 0; font-size: 14px;">Played Games</h3> <!-- Reduced font size -->
            <div class="stat-value">{played_games}</div>
            <p style="margin: 0; font-size: 12px;">{success_rate:.1f}% completion rate</p> <!-- Reduced font size -->
        </div>
        <div class="stat-card warning">
            <h3 style="margin: 0; font-size: 14px;">Remaining Games</h3> <!-- Reduced font size -->
            <div class="stat-value">{remaining_games}</div>
            <p style="margin: 0; font-size: 12px;">Games to play</p> <!-- Reduced font size -->
        </div>
    </div>
    
    
    <h2 style="font-size: 16px;">Game List</h2> <!-- Reduced font size -->
    <table>
        <thead>
            <tr>
                <th>SN</th>
                <th>Played</th>
                <th>Title</th>
                <th>Steam ID</th>
                <th>Genre</th>
                <th>Theme</th> <!-- NEW: Theme column -->
                <th>Description</th>
                <th class="mode-column">Game Mode</th>
                <th>Game Drive</th>
                <th>Original Title</th>
                <th>Resources</th>
            </tr>
        </thead>
        <tbody>
            {''.join(rows_html)}
        </tbody>
    </table>
    
    <div class="footer">
        Report generated by Game Manager • {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} • 
        Total: {total_games} games • Played: {played_games} • Remaining: {remaining_games}
    </div>
    
    <script>
        // Add simple interactivity
        document.addEventListener('DOMContentLoaded', function() {{
            // Add click confirmation for external links
            document.querySelectorAll('a[target="_blank"]').forEach(link => {{
                link.addEventListener('click', function(e) {{
                    if (!confirm('Open this link in a new tab?')) {{
                        e.preventDefault();
                    }}
                }});
            }});
            
            // Print button (optional)
            const printBtn = document.createElement('button');
            printBtn.textContent = '🖨️ Print Report';
            printBtn.style.cssText = 'position:fixed;bottom:15px;right:15px;padding:8px 12px;background:' + 
                                   '{COLOR_THEME["primary"]}' + ';color:white;border:none;border-radius:4px;cursor:pointer;font-size:12px;';
            printBtn.onclick = () => window.print();
            document.body.appendChild(printBtn);
        }});
    </script>
</body>
</html>"""
        
        # Write to file
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(html_content, encoding='utf-8')
        
        # Open in browser if requested
        if open_after:
            webbrowser.open(p.as_uri())
        
        return None
        
    except Exception as e:
        return f"Error exporting to HTML: {str(e)}"

# -----------------------------------------------------------------
# CONVENIENCE FUNCTION
# -----------------------------------------------------------------

def import_file_by_extension(path: str) -> Tuple[List[Dict], Optional[str]]:
    """Import file based on extension (.csv, .xlsx, .txt, .json, .db)."""
    ext = Path(path).suffix.lower()
    
    if ext == ".csv":
        return import_csv(path)
    elif ext in (".xlsx", ".xls"):
        return import_excel(path)
    elif ext in (".txt", ".list"):
        return import_txt(path)
    elif ext == ".json":
        return load_from_json(path)
    elif ext == ".db" or ext == ".sqlite":
        return load_from_sqlite(path)
    else:
        return [], f"Unsupported extension: {ext}"

# Continue from the export_games_to_pdf function...

def export_games_to_pdf_old(path: str, games: List[Dict], title: Optional[str] = None) -> Optional[str]:
    """
    Export games to a PDF file with formatted table.
    
    Args:
        path: Path to save the PDF file
        games: List of game dictionaries to export
        title: Title for the PDF document
    
    Returns:
        Error message or None if successful
    """
    # PDF configuration - adjust these values to fine-tune layout
    PDF_CONFIG = {
        "margin_mm": 4.0,      # Page margin in millimeters
        "font_size": 7.0,      # Base font size
        "scale": 1.00,         # Uniform column scale
        # Column width distribution (sum will be normalized)
        "col_percents": [
            0.015,  # SN
            0.020,  # Played
            0.08,   # Proper Title
            0.03,   # Steam ID
            0.05,   # Genre
            0.25,   # Description
            0.04,   # Game Mode
            0.03,   # Game Drive
            0.15,   # Original Title
            0.04    # Screenshots
        ]
    }
    
    try:
        # Try using ReportLab first (preferred)
        if REPORTLAB_AVAILABLE:
            page_size = landscape(A4)
            page_width_pts, page_height_pts = page_size
            margin_pts = PDF_CONFIG["margin_mm"] * mm
            usable_width = page_width_pts - (margin_pts * 2)
            
            # Setup styles
            styles = getSampleStyleSheet()
            title_style = styles["Title"]
            title_style.fontSize = max(12, PDF_CONFIG["font_size"] + 6)
            title_style.leading = title_style.fontSize + 2
            
            cell_style = ParagraphStyle(
                "cell",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=PDF_CONFIG["font_size"],
                leading=PDF_CONFIG["font_size"] + 1.2,
                spaceBefore=0,
                spaceAfter=0,
            )
            
            # Calculate column widths
            total_pct = sum(PDF_CONFIG["col_percents"]) or 1.0
            col_percents_norm = [p / total_pct for p in PDF_CONFIG["col_percents"]]
            base_widths = [usable_width * p for p in col_percents_norm]
            
            scaled_widths = [w * PDF_CONFIG["scale"] for w in base_widths]
            width_sum = sum(scaled_widths) or 1.0
            col_widths = [w * (usable_width / width_sum) for w in scaled_widths]
            
            # Prepare table data
            headers = ["SN", "Played", "Proper Title", "Steam ID", "Genre", "Description", 
                      "Game Mode", "Game Drive", "Original Title", "Screenshots"]
            table_data = [headers]
            
            total_games = len(games)
            played_count = 0
            
            for idx, game in enumerate(games, start=1):
                played = "Yes" if game.get("played") else "No"
                if game.get("played"):
                    played_count += 1
                
                # Get game data
                proper_title = str(game.get("title") or "")
                steam_id = str(game.get("app_id") or "")
                steam_link = game.get("steam_link") or ""
                
                # Create steam link if available
                steam_label = html.escape(steam_id) if steam_id else "Steam"
                if steam_link:
                    steam_cell_html = f'<a href="{html.escape(steam_link)}">{steam_label}</a>'
                else:
                    steam_cell_html = steam_label
                
                genre = str(game.get("genres") or "")
                description = str(game.get("description") or "").replace("\n", " ")
                game_mode = str(game.get("game_modes") or "")
                drive = str(game.get("game_drive") or "")
                original = str(game.get("original_title") or "")
                
                # Collect screenshot links
                screenshots = []
                for s in (game.get("screenshots") or []):
                    if s:
                        screenshots.append(s)
                
                for p in (game.get("image_cache_paths") or []):
                    if p and os.path.exists(p):
                        screenshots.append("file://" + os.path.abspath(p))
                    elif p:
                        screenshots.append(p)
                
                shot_links = []
                for s_idx, url in enumerate(screenshots, start=1):
                    safe_url = html.escape(url)
                    shot_links.append(f'<a href="{safe_url}">[{s_idx}]</a>')
                
                # Prepare cells with truncation
                plain_cells = [
                    str(idx),
                    played,
                    proper_title,
                    steam_label,
                    genre,
                    description,
                    game_mode,
                    drive,
                    original,
                    " ".join(f"[{n}]" for n in range(1, len(shot_links)+1))
                ]
                
                truncated_cells = []
                for col_i, plain in enumerate(plain_cells):
                    width_pts = col_widths[col_i]
                    truncated_plain = _truncate_to_two_lines(plain, width_pts, PDF_CONFIG["font_size"])
                    
                    # Rebuild HTML for special cells
                    if col_i == 3 and steam_link:
                        truncated_html = f'<a href="{html.escape(steam_link)}">{html.escape(truncated_plain)}</a>'
                        truncated_cells.append(Paragraph(truncated_html, cell_style))
                    elif col_i == 9 and shot_links:
                        # Limit number of shot links to fit two lines
                        avg_char_width = PDF_CONFIG["font_size"] * 0.55
                        chars_per_line = max(8, int(width_pts / avg_char_width))
                        max_links = max(0, int((chars_per_line * 2) / 4))
                        limited = shot_links[:max_links]
                        shots_html = " ".join(limited)
                        truncated_cells.append(Paragraph(shots_html, cell_style))
                    else:
                        truncated_cells.append(Paragraph(html.escape(truncated_plain), cell_style))
                
                table_data.append(truncated_cells)
            
            # Create table
            table = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
            table_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f2f2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), PDF_CONFIG["font_size"]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ])
            table.setStyle(table_style)
            
            # Build PDF document
            doc = SimpleDocTemplate(
                path,
                pagesize=page_size,
                leftMargin=margin_pts,
                rightMargin=margin_pts,
                topMargin=margin_pts,
                bottomMargin=margin_pts
            )
            
            story = []
            story.append(Paragraph(title or "Game Manager Export", title_style))
            story.append(Spacer(1, 6))
            story.append(table)
            story.append(Spacer(1, 8))
            
            # Add footer with statistics
            remaining = total_games - played_count
            footer_style = ParagraphStyle(
                "footer",
                parent=getSampleStyleSheet()["Normal"],
                fontSize=9,
                alignment=1
            )
            footer_text = f"Total games: {total_games}    Total played: {played_count}    Remaining: {remaining}"
            story.append(Paragraph(footer_text, footer_style))
            
            doc.build(story)
            return None
        
        # FPDF fallback
        if FPDF_AVAILABLE:
            pdf = FPDF(orientation="L", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=10)
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 8, title or "Game Manager Export", ln=True, align="C")
            pdf.ln(4)
            
            pdf.set_font("Arial", "B", 9)
            col_w = [10, 12, 60, 30, 30, 80, 30, 30, 40, 40]
            headers = ["SN", "Played", "Proper Title", "Steam ID", "Genre", "Description", 
                      "Game Mode", "Game Drive", "Original Title", "Screenshots"]
            
            # Draw header row
            for h, w in zip(headers, col_w):
                pdf.cell(w, 7, h, border=1)
            pdf.ln()
            
            pdf.set_font("Arial", size=8)
            total_games = len(games)
            played_count = 0
            
            # Draw data rows
            for i, game in enumerate(games, start=1):
                played = "Yes" if game.get("played") else "No"
                if game.get("played"):
                    played_count += 1
                
                title_text = (game.get("title") or "")[:60]
                steam_id = str(game.get("app_id") or "")[:40]
                steam_link = game.get("steam_link") or ""
                genre = (game.get("genres") or "")[:30]
                desc = (game.get("description") or "").replace("\n", " ")[:120]
                game_mode = (game.get("game_modes") or "")[:30]
                drive = (game.get("game_drive") or "")[:30]
                original = (game.get("original_title") or "")[:40]
                
                # Collect screenshots
                screenshots = []
                for s in (game.get("screenshots") or []):
                    if s:
                        screenshots.append(s)
                for p in (game.get("image_cache_paths") or []):
                    if p:
                        screenshots.append(p)
                
                shots_text = ", ".join(f"[{n}]" for n in range(1, len(screenshots)+1)) if screenshots else ""
                
                # Draw cells
                pdf.cell(col_w[0], 6, str(i), border=1)
                pdf.cell(col_w[1], 6, played, border=1)
                pdf.cell(col_w[2], 6, title_text, border=1)
                
                if steam_link:
                    pdf.cell(col_w[3], 6, steam_id or "Steam", border=1, link=steam_link)
                else:
                    pdf.cell(col_w[3], 6, steam_id, border=1)
                
                pdf.cell(col_w[4], 6, genre, border=1)
                pdf.cell(col_w[5], 6, desc, border=1)
                pdf.cell(col_w[6], 6, game_mode, border=1)
                pdf.cell(col_w[7], 6, drive, border=1)
                pdf.cell(col_w[8], 6, original, border=1)
                pdf.cell(col_w[9], 6, shots_text, border=1)
                pdf.ln()
            
            # Add footer
            remaining = total_games - played_count
            pdf.ln(4)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 6, f"Total games: {total_games}    Total played: {played_count}    Remaining: {remaining}", ln=True)
            
            pdf.output(path)
            return None
        
        return "No PDF library available. Install reportlab (preferred) or fpdf."
    
    except Exception as e:
        return f"Error exporting to PDF: {str(e)}"


def export_games_to_html_old(path: str, games: List[Dict], title: Optional[str] = None, open_after: bool = False) -> Optional[str]:
    """
    Export games to an HTML file with the requested columns and screenshot links.
    
    Args:
        path: Path to save the HTML file
        games: List of game dictionaries to export
        title: Title for the HTML document
        open_after: Open the generated HTML in the default browser
    
    Returns:
        Error message or None if successful
    """
    try:
        title_text = html.escape(title or "Game Manager Export")
        rows_html = []
        total = len(games)
        played_count = 0
        
        for i, game in enumerate(games, start=1):
            played = "Yes" if game.get("played") else "No"
            if game.get("played"):
                played_count += 1
            
            proper_title = html.escape(str(game.get("title") or ""))
            steam_id = html.escape(str(game.get("app_id") or ""))
            steam_link = game.get("steam_link") or ""
            
            if steam_link:
                steam_cell = f'<a href="{html.escape(steam_link)}" target="_blank">{steam_id or "Steam"}</a>'
            else:
                steam_cell = steam_id
            
            genre = html.escape(str(game.get("genres") or ""))
            desc = html.escape(str(game.get("description") or "")).replace("\n", "<br/>")
            game_mode = html.escape(str(game.get("game_modes") or ""))
            drive = html.escape(str(game.get("game_drive") or ""))
            original = html.escape(str(game.get("original_title") or ""))
            
            # Combine screenshots and cached paths
            screenshots = []
            for s in (game.get("screenshots") or []):
                if s:
                    screenshots.append(s)
            
            for p in (game.get("image_cache_paths") or []):
                if p:
                    # Convert local path to file:// if exists
                    if Path(p).exists():
                        screenshots.append("file://" + str(Path(p).resolve()))
                    else:
                        screenshots.append(p)
            
            shot_links = []
            for idx_s, url in enumerate(screenshots, start=1):
                safe_url = html.escape(url)
                shot_links.append(f'<a href="{safe_url}" target="_blank">[{idx_s}]</a>')
            
            shots_cell = " ".join(shot_links)
            
            rows_html.append(
                "<tr>"
                f"<td>{i}</td>"
                f"<td>{played}</td>"
                f"<td>{proper_title}</td>"
                f"<td>{steam_cell}</td>"
                f"<td>{genre}</td>"
                f"<td>{desc}</td>"
                f"<td>{game_mode}</td>"
                f"<td>{drive}</td>"
                f"<td>{original}</td>"
                f"<td>{shots_cell}</td>"
                "</tr>"
            )
        
        footer_total = f"Total games: {total} &nbsp;&nbsp; Total played: {played_count} &nbsp;&nbsp; Remaining: {total - played_count}"
        
        html_doc = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8"/>
<title>{title_text}</title>
<style>
body {{ font-family: Arial, Helvetica, sans-serif; font-size: 11px; color: #222; margin: 20px; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 10px; }}
th, td {{ border: 1px solid #ddd; padding: 6px; vertical-align: top; text-align: left; }}
th {{ background: #f4f4f4; font-weight: bold; }}
tr:nth-child(even) {{ background-color: #f9f9f9; }}
tr:hover {{ background-color: #f0f0f0; }}
td a {{ color: #1a0dab; text-decoration: none; }}
td a:hover {{ text-decoration: underline; }}
.footer {{ margin-top: 15px; font-weight: bold; padding: 10px; background-color: #f4f4f4; border-radius: 4px; }}
h2 {{ color: #333; border-bottom: 2px solid #4CAF50; padding-bottom: 5px; }}
</style>
</head>
<body>
<h2>{title_text}</h2>
<table>
<thead>
<tr>
<th>SN</th><th>Played</th><th>Proper Title</th><th>Steam ID</th><th>Genre</th><th>Description</th>
<th>Game Mode</th><th>Game Drive</th><th>Original Title</th><th>Screenshots</th>
</tr>
</thead>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
<div class="footer">{footer_total}</div>
</body>
</html>
"""
        
        # Write file
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(html_doc, encoding="utf-8")
        
        # Open in browser if requested
        if open_after:
            webbrowser.open(p.as_uri())
        
        return None
    
    except Exception as e:
        return f"Error exporting to HTML: {str(e)}"


def export_games_to_csv(path: str, games: List[Dict]) -> Optional[str]:
    """
    Export games to a CSV file.
    
    Args:
        path: Path to save the CSV file
        games: List of game dictionaries to export
    
    Returns:
        Error message or None if successful
    """
    try:
        # Define field order
        fields = [
            "title", "app_id", "release_date", "developer", "publisher",
            "genres", "description", "cover_url", "trailer_webm",
            "steam_link", "steamdb_link", "pcgw_link", "igdb_link",
            "game_drive", "scene_repack", "game_modes", "original_title",
            "original_title_base", "original_title_version", "original_notes",
            "patch_version", "player_perspective", "themes", "igdb_id",
            "played", "save_location", "shortcut_links"
        ]
        
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            
            for game in games:
                row = {}
                for field in fields:
                    value = game.get(field)
                    
                    if isinstance(value, list):
                        # Join lists with pipe separator
                        row[field] = "|".join(str(item) for item in value if item)
                    elif isinstance(value, bool):
                        row[field] = "Yes" if value else "No"
                    else:
                        row[field] = str(value or "")
                
                writer.writerow(row)
        
        return None
    
    except Exception as e:
        return f"Error exporting to CSV: {str(e)}"


def export_games_to_excel(path: str, games: List[Dict]) -> Optional[str]:
    """
    Export games to an Excel file.
    
    Args:
        path: Path to save the Excel file
        games: List of game dictionaries to export
    
    Returns:
        Error message or None if successful
    """
    try:
        if not PANDAS_AVAILABLE:
            return "pandas library not available for Excel export"
        
        # Prepare data for DataFrame
        data = []
        fields = [
            "title", "app_id", "release_date", "developer", "publisher",
            "genres", "description", "cover_url", "trailer_webm",
            "steam_link", "steamdb_link", "pcgw_link", "igdb_link",
            "game_drive", "scene_repack", "game_modes", "original_title",
            "original_title_base", "original_title_version", "original_notes",
            "patch_version", "player_perspective", "themes", "igdb_id",
            "played", "save_location", "shortcut_links",
            "screenshots", "image_cache_paths", "savegame_location"
        ]
        
        for game in games:
            row = {}
            for field in fields:
                value = game.get(field)
                
                if isinstance(value, list):
                    # Join lists with pipe separator
                    row[field] = "|".join(str(item) for item in value if item)
                elif isinstance(value, bool):
                    row[field] = "Yes" if value else "No"
                else:
                    row[field] = str(value or "")
            
            data.append(row)
        
        # Create DataFrame and save to Excel
        df = pd.DataFrame(data, columns=fields)
        
        # Auto-adjust column widths
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Games', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Games']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return None
    
    except Exception as e:
        return f"Error exporting to Excel: {str(e)}"


# --------------------------------------------------------------------------
# File Type Detection and Auto-Import
# --------------------------------------------------------------------------

def import_file_by_extension(path: str) -> Tuple[List[Dict], Optional[str]]:
    """
    Auto-detect file type by extension and call the appropriate importer.
    
    Args:
        path: Path to the file to import
    
    Returns:
        Tuple of (list of game dictionaries, error message or None)
    """
    ext = Path(path).suffix.lower()
    
    if ext in (".csv",):
        return import_csv(path)
    elif ext in (".xlsx", ".xls"):
        return import_excel(path)
    elif ext in (".txt", ".list"):
        return import_txt(path)
    elif ext in (".json",):
        return load_from_json(path)
    elif ext in (".db", ".sqlite", ".sqlite3"):
        return load_from_sqlite(path)
    else:
        return [], f"Unsupported file extension: {ext}"


def export_file_by_extension(path: str, games: List[Dict], **kwargs) -> Optional[str]:
    """
    Auto-detect file type by extension and call the appropriate exporter.
    
    Args:
        path: Path to save the file
        games: List of game dictionaries to export
        **kwargs: Additional arguments for specific exporters
    
    Returns:
        Error message or None if successful
    """
    ext = Path(path).suffix.lower()
    
    if ext in (".csv",):
        return export_games_to_csv(path, games)
    elif ext in (".xlsx", ".xls"):
        return export_games_to_excel(path, games)
    elif ext in (".pdf",):
        title = kwargs.get('title', None)
        return export_games_to_pdf(path, games, title)
    elif ext in (".html", ".htm"):
        title = kwargs.get('title', None)
        open_after = kwargs.get('open_after', False)
        return export_games_to_html(path, games, title, open_after)
    elif ext in (".json",):
        return save_to_json(path, games)
    elif ext in (".db", ".sqlite", ".sqlite3"):
        return save_to_sqlite(path, games)
    else:
        return f"Unsupported file extension: {ext}"


# --------------------------------------------------------------------------
# Enhanced Game Processing with PCGW Integration
# --------------------------------------------------------------------------

def enhance_games_with_pcgw(games: List[Dict], max_workers: int = 3, force_refresh: bool = False) -> List[Dict]:
    """
    Enhance game list with PCGW information (save locations, etc.).
    
    Args:
        games: List of game dictionaries to enhance
        max_workers: Maximum concurrent workers for scraping
        force_refresh: Force fresh scrape (skip cache)
    
    Returns:
        Enhanced game list
    """
    try:
        # Check if PCGW scraper is available
        from utils_sanitize import update_game_with_pcgw_info
    except ImportError:
        print("PCGW scraper not available. Skipping enhancement.")
        return games
    
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    enhanced_games = games.copy()
    
    def enhance_single_game(game):
        """Enhance a single game with PCGW info."""
        try:
            return update_game_with_pcgw_info(game, force_refresh)
        except Exception as e:
            print(f"Error enhancing game {game.get('title', 'Unknown')}: {e}")
            return game  # Return original on error
    
    # Use thread pool for concurrent enhancement
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(enhance_single_game, game): i 
                  for i, game in enumerate(enhanced_games)}
        
        for future in as_completed(futures):
            idx = futures[future]
            try:
                enhanced_games[idx] = future.result()
            except Exception as e:
                print(f"Error processing game at index {idx}: {e}")
    
    return enhanced_games


def batch_update_save_locations(games: List[Dict], output_file: Optional[str] = None) -> List[Dict]:
    """
    Batch update save locations for all games using PCGW.
    
    Args:
        games: List of game dictionaries
        output_file: Optional file to save updated games
    
    Returns:
        Updated game list
    """
    print(f"Updating save locations for {len(games)} games...")
    
    updated_games = enhance_games_with_pcgw(games)
    
    # Count statistics
    games_with_saves = sum(1 for g in updated_games if g.get("savegame_location"))
    total_saves = sum(len(g.get("savegame_location", [])) for g in updated_games)
    
    print(f"Updated {games_with_saves}/{len(games)} games with save locations")
    print(f"Total save locations found: {total_saves}")
    
    # Save to file if requested
    if output_file:
        error = save_to_json(output_file, updated_games)
        if error:
            print(f"Error saving to {output_file}: {error}")
        else:
            print(f"Saved updated games to {output_file}")
    
    return updated_games


# --------------------------------------------------------------------------
# Data Validation and Cleaning
# --------------------------------------------------------------------------

def validate_game_data(game: Dict) -> Dict[str, List[str]]:
    """
    Validate a game dictionary and return validation errors.
    
    Args:
        game: Game dictionary to validate
    
    Returns:
        Dictionary of validation errors by field
    """
    errors = {}
    
    # Required fields
    if not game.get("title"):
        errors.setdefault("title", []).append("Title is required")
    
    # Field length validation
    if game.get("description") and len(game["description"]) > 5000:
        errors.setdefault("description", []).append("Description too long (max 5000 chars)")
    
    # URL validation
    url_fields = ["steam_link", "steamdb_link", "pcgw_link", "igdb_link", "cover_url", "trailer_webm"]
    for field in url_fields:
        value = game.get(field)
        if value and not value.startswith(("http://", "https://")):
            errors.setdefault(field, []).append("Invalid URL format")
    
    # Date validation (basic)
    release_date = game.get("release_date")
    if release_date:
        # Simple date format check (YYYY-MM-DD)
        import re
        if not re.match(r'^\d{4}(-\d{2}){0,2}$', release_date):
            errors.setdefault("release_date", []).append("Invalid date format (use YYYY-MM-DD)")
    
    # List field validation
    list_fields = ["screenshots", "image_cache_paths", "savegame_location"]
    for field in list_fields:
        value = game.get(field)
        if value and not isinstance(value, list):
            errors.setdefault(field, []).append("Must be a list")
    
    return errors


def clean_game_data(game: Dict) -> Dict:
    """
    Clean and normalize game data.
    
    Args:
        game: Game dictionary to clean
    
    Returns:
        Cleaned game dictionary
    """
    cleaned = empty_game()
    cleaned.update(game)
    
    # Trim string fields
    string_fields = ["title", "app_id", "release_date", "developer", "publisher",
                    "genres", "description", "cover_url", "trailer_webm",
                    "steam_link", "steamdb_link", "pcgw_link", "igdb_link",
                    "game_drive", "scene_repack", "game_modes", "original_title",
                    "original_title_base", "original_title_version", "original_notes",
                    "patch_version", "player_perspective", "themes", "igdb_id",
                    "save_location", "shortcut_links"]
    
    for field in string_fields:
        if field in cleaned and cleaned[field] is not None:
            cleaned[field] = str(cleaned[field]).strip()
    
    # Ensure list fields are lists
    list_fields = ["screenshots", "image_cache_paths", "savegame_location"]
    for field in list_fields:
        value = cleaned.get(field)
        if value is None:
            cleaned[field] = []
        elif isinstance(value, str):
            # Split pipe-separated strings
            cleaned[field] = [item.strip() for item in value.split("|") if item.strip()]
        elif not isinstance(value, list):
            cleaned[field] = []
    
    # Normalize played field
    played = cleaned.get("played")
    if isinstance(played, str):
        cleaned["played"] = played.lower() in ("yes", "y", "true", "1", "checked")
    elif not isinstance(played, bool):
        cleaned["played"] = False
    
    # Generate missing app_id if possible
    if not cleaned["app_id"] and cleaned["steam_link"]:
        # Extract app_id from Steam URL
        import re
        match = re.search(r'/app/(\d+)', cleaned["steam_link"])
        if match:
            cleaned["app_id"] = match.group(1)
    
    # Ensure title is populated
    if not cleaned["title"] and cleaned["original_title"]:
        cleaned["title"] = cleaned["original_title"]
    
    return cleaned


def deduplicate_games(games: List[Dict], key_field: str = "app_id") -> List[Dict]:
    """
    Remove duplicate games based on a key field.
    
    Args:
        games: List of game dictionaries
        key_field: Field to use for deduplication
    
    Returns:
        Deduplicated game list
    """
    seen = set()
    deduplicated = []
    
    for game in games:
        key = game.get(key_field)
        if not key:
            # If no key, use title as fallback
            key = game.get("title", "")
        
        if key not in seen:
            seen.add(key)
            deduplicated.append(game)
        else:
            print(f"Duplicate found: {game.get('title', 'Unknown')}")
    
    return deduplicated


# --------------------------------------------------------------------------
# Statistics and Reporting
# --------------------------------------------------------------------------

def get_game_statistics(games: List[Dict]) -> Dict[str, Any]:
    """
    Generate statistics from game list.
    
    Args:
        games: List of game dictionaries
    
    Returns:
        Dictionary of statistics
    """
    total = len(games)
    
    stats = {
        "total_games": total,
        "played_games": sum(1 for g in games if g.get("played")),
        "games_with_screenshots": sum(1 for g in games if g.get("screenshots")),
        "games_with_save_locations": sum(1 for g in games if g.get("savegame_location")),
        "games_with_app_id": sum(1 for g in games if g.get("app_id")),
        "games_by_genre": {},
        "games_by_developer": {},
        "games_by_year": {}
    }
    
    # Genre distribution
    for game in games:
        genres = game.get("genres", "")
        if genres:
            for genre in genres.split(","):
                genre = genre.strip()
                if genre:
                    stats["games_by_genre"][genre] = stats["games_by_genre"].get(genre, 0) + 1
    
    # Developer distribution
    for game in games:
        developer = game.get("developer", "")
        if developer:
            stats["games_by_developer"][developer] = stats["games_by_developer"].get(developer, 0) + 1
    
    # Release year distribution
    for game in games:
        release_date = game.get("release_date", "")
        if release_date and len(release_date) >= 4:
            year = release_date[:4]
            stats["games_by_year"][year] = stats["games_by_year"].get(year, 0) + 1
    
    return stats


def print_statistics(games: List[Dict]):
    """
    Print formatted statistics for game list.
    
    Args:
        games: List of game dictionaries
    """
    stats = get_game_statistics(games)
    
    print("\n" + "="*50)
    print("GAME STATISTICS")
    print("="*50)
    print(f"Total games: {stats['total_games']}")
    print(f"Played: {stats['played_games']} ({stats['played_games']/stats['total_games']*100:.1f}%)")
    print(f"With screenshots: {stats['games_with_screenshots']}")
    print(f"With save locations: {stats['games_with_save_locations']}")
    print(f"With Steam App ID: {stats['games_with_app_id']}")
    
    # Top genres
    if stats["games_by_genre"]:
        print("\nTop 5 Genres:")
        top_genres = sorted(stats["games_by_genre"].items(), key=lambda x: x[1], reverse=True)[:5]
        for genre, count in top_genres:
            print(f"  {genre}: {count}")
    
    # Recent releases
    if stats["games_by_year"]:
        print("\nReleases by Year:")
        recent_years = sorted(stats["games_by_year"].items(), key=lambda x: x[0], reverse=True)[:5]
        for year, count in recent_years:
            print(f"  {year}: {count}")


# --------------------------------------------------------------------------
# CLI Interface for Testing
# --------------------------------------------------------------------------

def test_cli():
    """Simple CLI for testing import/export functionality."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python import_export.py <command> [options]")
        print("\nCommands:")
        print("  import <file>          Import games from file")
        print("  export <file>          Export games to file")
        print("  stats <file>           Show statistics for games file")
        print("  validate <file>        Validate games in file")
        print("  clean <in> <out>       Clean and normalize games")
        print("  enhancesaves <in> <out> Enhance with PCGW save locations")
        print("  merge <base> <new> <out> Merge two game files")
        return
    
    command = sys.argv[1].lower()
    
    if command == "import" and len(sys.argv) >= 3:
        input_file = sys.argv[2]
        games, error = import_file_by_extension(input_file)
        
        if error:
            print(f"Error: {error}")
        else:
            print(f"Successfully imported {len(games)} games from {input_file}")
            if games:
                print(f"First game: {games[0].get('title', 'Unknown')}")
    
    elif command == "export" and len(sys.argv) >= 4:
        input_file = sys.argv[2]
        output_file = sys.argv[3]
        
        # Load games
        games, error = import_file_by_extension(input_file)
        if error:
            print(f"Error loading: {error}")
            return
        
        # Export games
        error = export_file_by_extension(output_file, games)
        if error:
            print(f"Error exporting: {error}")
        else:
            print(f"Successfully exported {len(games)} games to {output_file}")
    
    elif command == "stats" and len(sys.argv) >= 3:
        input_file = sys.argv[2]
        games, error = import_file_by_extension(input_file)
        
        if error:
            print(f"Error: {error}")
        else:
            print_statistics(games)
    
    elif command == "validate" and len(sys.argv) >= 3:
        input_file = sys.argv[2]
        games, error = import_file_by_extension(input_file)
        
        if error:
            print(f"Error: {error}")
        else:
            print(f"Validating {len(games)} games...")
            error_count = 0
            
            for i, game in enumerate(games):
                errors = validate_game_data(game)
                if errors:
                    error_count += 1
                    print(f"\nGame {i+1}: {game.get('title', 'Unknown')}")
                    for field, field_errors in errors.items():
                        print(f"  {field}: {', '.join(field_errors)}")
            
            if error_count == 0:
                print("All games are valid!")
            else:
                print(f"\nFound errors in {error_count} games")
    
    elif command == "clean" and len(sys.argv) >= 4:
        input_file = sys.argv[2]
        output_file = sys.argv[3]
        
        games, error = import_file_by_extension(input_file)
        if error:
            print(f"Error loading: {error}")
            return
        
        # Clean games
        cleaned_games = [clean_game_data(g) for g in games]
        
        # Save cleaned games
        error = save_to_json(output_file, cleaned_games)
        if error:
            print(f"Error saving: {error}")
        else:
            print(f"Cleaned {len(cleaned_games)} games, saved to {output_file}")
    
    elif command == "enhancesaves" and len(sys.argv) >= 4:
        input_file = sys.argv[2]
        output_file = sys.argv[3]
        
        games, error = import_file_by_extension(input_file)
        if error:
            print(f"Error loading: {error}")
            return
        
        # Enhance with save locations
        enhanced_games = batch_update_save_locations(games, output_file)
        print(f"Enhanced {len(enhanced_games)} games with PCGW save locations")
    
    elif command == "merge" and len(sys.argv) >= 5:
        base_file = sys.argv[2]
        new_file = sys.argv[3]
        output_file = sys.argv[4]
        
        # Load both files
        base_games, error1 = import_file_by_extension(base_file)
        new_games, error2 = import_file_by_extension(new_file)
        
        if error1 or error2:
            print(f"Error loading files: {error1 or error2}")
            return
        
        # Merge games
        merged_games = merge_imported_rows(base_games, new_games, prefer_imported=True)
        
        # Save merged games
        error = save_to_json(output_file, merged_games)
        if error:
            print(f"Error saving: {error}")
        else:
            print(f"Merged {len(base_games)} + {len(new_games)} = {len(merged_games)} games")
            print(f"Saved to {output_file}")
    
    else:
        print(f"Unknown command: {command}")
        test_cli()


# --------------------------------------------------------------------------
# Main Execution Block
# --------------------------------------------------------------------------

if __name__ == "__main__":
    # Run CLI if script is executed directly
    test_cli()